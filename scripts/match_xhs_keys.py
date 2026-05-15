#!/usr/bin/env python3
"""XHS 数据 Key 匹配工具：将 Excel 笔记标题与 Notion 数据库中的 key 关联。

用法:
  python scripts/match_xhs_keys.py ~/Data/笔记列表明细表_20260507-20260514.xlsx
  python scripts/match_xhs_keys.py ~/Data/笔记列表明细表_20260507-20260514.xlsx --threshold 0.6

Excel 要求: 第 A 列为空 Key 列，第 B 列为笔记标题。
输出: 原地写入 A 列 Key，未匹配的留空。
"""

import openpyxl, os, sys, time, argparse
from difflib import SequenceMatcher

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def parse_date_range(filename: str) -> tuple[str, str] | None:
    """从文件名提取日期范围，如 笔记列表明细表_20260507-20260514 → ('2026.05.07', '2026.05.14')"""
    import re
    m = re.search(r'(\d{8})-(\d{8})', filename)
    if m:
        start = f"{m.group(1)[:4]}.{m.group(1)[4:6]}.{m.group(1)[6:8]}"
        end = f"{m.group(2)[:4]}.{m.group(2)[4:6]}.{m.group(2)[6:8]}"
        return start, end
    return None


def load_notion_entries(date_start: str = "", date_end: str = "") -> list[tuple[str, str]]:
    """从 Notion 加载发布XHS=True 的条目，返回 [(title, key), ...]"""
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

    API_KEY = os.environ["NOTION_API_KEY"]
    DB_ID = os.environ["NOTION_DATABASE_ID"]
    HEADERS = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
        "Notion-Version": "2022-06-28",
    }

    all_pages = []
    has_more, cursor = True, None

    # 构建 Notion API 复合 filter
    filters = [{"property": "发布XHS", "checkbox": {"equals": True}}]
    if date_start and date_end:
        filters.append({
            "property": "发布XHS时间",
            "date": {"on_or_after": date_start.replace(".", "-"), "on_or_before": date_end.replace(".", "-")}
        })

    and_filter = {"and": filters}
    print(f"📡 查询 Notion (发布XHS=True, 日期 {date_start}~{date_end})..." if date_start
          else "📡 查询 Notion (发布XHS=True)...")

    while has_more:
        payload: dict = {"filter": and_filter, "page_size": 100}
        if cursor:
            payload["start_cursor"] = cursor

        resp = __import__('requests').post(
            f"https://api.notion.com/v1/databases/{DB_ID}/query",
            headers=HEADERS, json=payload, timeout=20,
        )
        data = resp.json()
        all_pages.extend(data.get("results", []))
        has_more = data.get("has_more", False)
        cursor = data.get("next_cursor")
        if has_more:
            time.sleep(0.3)

    entries = []
    for page in all_pages:
        props = page.get("properties", {})
        title = "".join(t.get("plain_text", "") for t in props.get("Name", {}).get("title", [])).strip()
        key = "".join(r.get("plain_text", "") for r in props.get("key", {}).get("rich_text", []))
        if title:
            entries.append((title, key))

    return entries


def main():
    parser = argparse.ArgumentParser(description="XHS Key 匹配工具")
    parser.add_argument("excel", help="Excel 文件路径（第 A 列 Key, 第 B 列标题）")
    parser.add_argument("--threshold", "-t", type=float, default=0.5,
                        help="模糊匹配相似度阈值（默认 0.5，范围 0-1）")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument("--col-title", type=int, default=2, help="标题列（默认 B=2）")
    parser.add_argument("--col-key", type=int, default=1, help="Key 列（默认 A=1）")
    parser.add_argument("--start-row", type=int, default=3, help="数据起始行（默认 3）")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)

    # ── 读取 Excel ──
    print(f"📂 {os.path.basename(args.excel)}")
    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    excel_data = []
    for row_idx in range(args.start_row, ws.max_row + 1):
        title = ws.cell(row=row_idx, column=args.col_title).value
        if title:
            excel_data.append((row_idx, str(title).strip()))
    print(f"   Excel: {len(excel_data)} 条笔记")

    # ── 自动推断日期范围 ──
    date_range = parse_date_range(os.path.basename(args.excel))
    date_start, date_end = date_range if date_range else ("", "")

    # ── 加载 Notion ──
    notion_entries = load_notion_entries(date_start, date_end)
    print(f"   Notion: {len(notion_entries)} 条 ({date_start} ~ {date_end})" if date_start
          else f"   Notion: {len(notion_entries)} 条（全部）")

    # ── 匹配 ──
    matched = 0
    low_conf = []

    for row_idx, excel_title in excel_data:
        best_score = 0.0
        best_key = ""
        best_ntitle = ""

        for ntitle, nkey in notion_entries:
            score = SequenceMatcher(None, excel_title, ntitle).ratio()
            if score > best_score:
                best_score = score
                best_key = nkey
                best_ntitle = ntitle

        if best_score >= args.threshold:
            ws.cell(row=row_idx, column=args.col_key).value = best_key
            matched += 1
        else:
            low_conf.append((row_idx, excel_title, best_score, best_ntitle))

    # ── 保存 ──
    wb.save(args.excel)
    print(f"\n✅ 匹配: {matched}/{len(excel_data)} (阈值 {args.threshold})")

    if low_conf:
        print(f"\n⚠️  {len(low_conf)} 条未匹配（需手动）:")
        for row_idx, title, score, best in low_conf:
            print(f"  [行{row_idx:2d}] score={score:.2f} | {title[:55]}")
            if best:
                print(f"         候选: {best[:60]}")


if __name__ == "__main__":
    main()
